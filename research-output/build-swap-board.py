"""
build-swap-board.py - The Nasties LIVE SWAP BOARD.

A slot-ladder auction board built for live use: names grouped by the roster
slot they fill, best value first, each with Win $ (target bid) and Walk $ (hard
ceiling). Cross a name off (type GONE) the moment he leaves the board or you
pass; the top name still open in that bucket is always your next move. A budget
helper up top does the running math so it does not have to live in your head.

Every price traces to research-output/dataset.json (valueBand + land odds) from
the same pull behind the target sheet. No fabrication. Run after the target
sheet; regenerate when the data pull changes.

Writes {League}_Live_Swap_Board_{date}.xlsx to research-output/ AND ~/Downloads/.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DS = os.path.join(HERE, "dataset.json")

with open(DS, encoding="utf-8") as f:
    ds = json.load(f)
players = ds if isinstance(ds, list) else (ds.get("players") or ds.get("rankings") or next(v for v in ds.values() if isinstance(v, list)))
meta = ds.get("meta", {}) if isinstance(ds, dict) else {}
league = (meta.get("league") or "The Nasties").replace("The ", "").strip() or "Nasties"
date = (meta.get("generatedAt") or "")[:10] or "latest"

by = {}
for p in players:
    nm = (p.get("name") or "").strip()
    if nm and nm not in by:
        by[nm] = p

def band(nm):
    p = by.get(nm) or {}
    vb = p.get("valueBand") or {}
    return vb.get("low"), vb.get("base"), vb.get("high")

def land(nm):
    p = by.get(nm) or {}
    for k in ("landProbability", "landOdds", "landProb"):
        if p.get(k) is not None:
            return round(p[k] * 100)
    return None

def row(nm, pos, note=""):
    lo, base, hi = band(nm)
    lp = land(nm)
    return ["", nm, pos, base, hi, (f"{lp}%" if lp is not None else "-"), note]

# Buckets: (title, subtitle, [rows]). Order = draft priority, top bucket first.
BUCKETS = [
    ("1. ANCHOR", "Win exactly ONE. Pay up. If sniped past Walk, your anchor becomes a WR from bucket 2.", [
        row("Jahmyr Gibbs", "RB", "The one seat you pay full freight for. Lands 92% when you commit."),
    ]),
    ("2. BIG BUY #2", "Win exactly ONE. Interchangeable - take whichever falls under its Walk. Never chase a specific name.", [
        row("Jaxon Smith-Njigba", "WR", "Best-graded Gibbs pair (9.0-5.0)"),
        row("Amon-Ra St. Brown", "WR", "Best-graded Gibbs pair (9.0-5.0), cheaper than JSN"),
        row("Puka Nacua", "WR", "Questionable tag, priced at full room"),
        row("Bijan Robinson", "RB", "Only for RB-RB or if Gibbs is lost"),
    ]),
    ("3. RB DEPTH", "Win 2-3. RB runs COOL (0.8x) = your RELIABLE cheap zone. Work down as names go.", [
        row("Quinshon Judkins", "RB", "+$7 pocket, top RB2 target"),
        row("Cam Skattebo", "RB", "breakout dart"),
        row("Tony Pollard", "RB", ""),
        row("J.K. Dobbins", "RB", "fragile"),
        row("Zach Charbonnet", "RB", "PUP - stash only, trim the bid"),
        row("Jordan Mason", "RB", "$1 body"),
    ]),
    ("4. WR DEPTH", "Win 2-3. WR runs HOT (1.2x) = CONTESTED. If bid over Win, fall to the next name, do not chase.", [
        row("Garrett Wilson", "WR", "+$4 pocket"),
        row("Zay Flowers", "WR", "+$5 pocket, Questionable"),
        row("Rome Odunze", "WR", "+$4 pocket"),
        row("Carnell Tate", "WR", "deep sleeper"),
        row("Courtland Sutton", "WR", "+$5 pocket, $1 floor"),
        row("Josh Downs", "WR", "$1 dart"),
        row("Wan'Dale Robinson", "WR", "$1 dart"),
    ]),
    ("5. QB", "Win ONE. Ladder - take the first that falls under its Win. Budget up to ~$17, do NOT bank on $1 Nix.", [
        row("Bo Nix", "QB", "+$10, biggest QB bargain"),
        row("Jaxson Dart", "QB", "+$5, cheap upside"),
        row("Jayden Daniels", "QB", ""),
        row("Jalen Hurts", "QB", "+$6, real weekly starter"),
        row("Kyler Murray", "QB", "punt option"),
    ]),
    ("6. TE", "Win ONE. Kittle is the value; pay-up options only if the WR anchor came cheap.", [
        row("George Kittle", "TE", "+$5, best value, fragile"),
        row("Dallas Goedert", "TE", "$1-2 dart"),
        row("Harold Fannin Jr.", "TE", "cheap upside"),
        row("Trey McBride", "TE", "pay-up option"),
        row("Brock Bowers", "TE", "pay-up option, near full price"),
    ]),
    ("7. DEF", "Win ONE, last, $1-3 MAX. A defense is a stream slot - never pay up. Cheapest open name wins.", [
        row("Pittsburgh Steelers", "DEF", "$1-3, stream it"),
        row("Denver Broncos", "DEF", "$1-3, stream it"),
        row("Houston Texans", "DEF", "$1-3, stream it"),
        row("Philadelphia Eagles", "DEF", "$1-3, stream it"),
    ]),
]

NAVY = "1F2A44"; GOLD = "C8A24B"; WHITE = "FFFFFF"
hdr_fill = PatternFill("solid", fgColor=NAVY)
bucket_fill = PatternFill("solid", fgColor=GOLD)
win_fill = PatternFill("solid", fgColor="FFF2CC")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "Live Swap Board"
ws.sheet_view.showGridLines = False

WIDTHS = [11, 26, 6, 8, 8, 8, 40]
for ci, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

r = 1
# Title
ws.cell(r, 1, f"{league.upper()} - LIVE SWAP BOARD  ({date})").font = Font(bold=True, size=15, color=GOLD)
r += 1
ws.cell(r, 1, "Draft top bucket down. Type GONE the second a name leaves the board or you pass. Your next move is always the top OPEN name. Never bid past Walk $.").font = Font(size=10, italic=True)
r += 2

# Budget helper
ws.cell(r, 1, "BUDGET HELPER").font = Font(bold=True, size=11, color=WHITE)
for c in range(1, 8):
    ws.cell(r, c).fill = hdr_fill
r += 1
help_rows = [
    ("Total budget", 200, "fixed"),
    ("Spent so far", 0, "<- type this as you go"),
    ("Remaining", "=B{sp}-B{spent}", "auto"),
    ("Roster slots total", 13, "QB1/RB1/WR1/TE1/FLEX3/DEF1/Bench5"),
    ("Slots still open", 13, "<- type this as you go"),
    ("Max avg left / open slot", "=IF(B{open}>0,ROUND(B{rem}/B{open},1),0)", "auto - keep at least $1 x open darts in pocket"),
]
base_r = r
labels_at = {}
for i, (lbl, val, note) in enumerate(help_rows):
    rr = base_r + i
    labels_at[lbl] = rr
for i, (lbl, val, note) in enumerate(help_rows):
    rr = base_r + i
    ws.cell(rr, 1, lbl).font = Font(bold=True, size=10)
    if isinstance(val, str) and val.startswith("="):
        formula = val.format(
            sp=labels_at["Total budget"], spent=labels_at["Spent so far"],
            rem=labels_at["Remaining"], open=labels_at["Slots still open"],
        )
        ws.cell(rr, 2, formula).font = Font(bold=True, size=11, color="107C10")
    else:
        ws.cell(rr, 2, val).font = Font(size=11)
    ws.cell(rr, 2).fill = win_fill
    ws.cell(rr, 3, note).font = Font(size=9, italic=True, color="666666")
r = base_r + len(help_rows) + 1

# Budget-frame reminder line
ws.cell(r, 1, "PLAN: two studs cost ~$155-160. Hold ~$5. That leaves ~$40 for the other 11 slots. Depth is the whole game - lean on the RB tier (cool), let the hot WRs go.").font = Font(size=10, italic=True, color="C00000")
r += 2

HEAD = ["Got/Gone", "Player", "Pos", "Win $", "Walk $", "Land", "Note"]
status_rows = []  # for conditional formatting + validation

for title, subtitle, rows in BUCKETS:
    # Bucket header band
    ws.cell(r, 1, title).font = Font(bold=True, size=12, color=NAVY)
    for c in range(1, 8):
        ws.cell(r, c).fill = bucket_fill
    ws.cell(r, 2).value = None
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, subtitle).font = Font(size=9, italic=True, color=NAVY)
    ws.cell(r, 2).alignment = left
    ws.row_dimensions[r].height = 26
    r += 1
    # Column header
    for ci, h in enumerate(HEAD, 1):
        c = ws.cell(r, ci, h); c.fill = hdr_fill; c.font = Font(bold=True, color=WHITE, size=10); c.alignment = center; c.border = border
    r += 1
    for data in rows:
        for ci, val in enumerate(data, 1):
            c = ws.cell(r, ci, val); c.border = border
            c.alignment = left if ci in (2, 7) else center
            if ci == 4:  # Win col
                c.fill = win_fill; c.font = Font(bold=True)
        status_rows.append(r)
        r += 1
    r += 1  # gap between buckets

# Status dropdown + conditional formatting (GONE = gray + strike, GOT = green)
dv = DataValidation(type="list", formula1='"GOT,GONE"', allow_blank=True)
ws.add_data_validation(dv)
first, last = min(status_rows), max(status_rows)
dv.add(f"A{first}:A{last}")

gone_style = DifferentialStyle(font=Font(strike=True, color="999999"), fill=PatternFill("solid", fgColor="EEEEEE"))
got_style = DifferentialStyle(font=Font(bold=True, color="107C10"), fill=PatternFill("solid", fgColor="E2EFDA"))
rng = f"A{first}:G{last}"
ws.conditional_formatting.add(rng, Rule(type="expression", formula=[f'$A{first}="GONE"'], dxf=gone_style, stopIfTrue=False))
ws.conditional_formatting.add(rng, Rule(type="expression", formula=[f'$A{first}="GOT"'], dxf=got_style, stopIfTrue=False))

ws.freeze_panes = "A" + str(base_r + len(help_rows) + 3)

out_names = [
    os.path.join(HERE, f"{league}_Live_Swap_Board_{date}.xlsx"),
    os.path.join(os.path.expanduser("~"), "Downloads", f"{league}_Live_Swap_Board_{date}.xlsx"),
]
for out in out_names:
    wb.save(out)
    print("SAVED:", out, os.path.getsize(out), "bytes")
print("buckets:", len(BUCKETS), "| player rows:", len(status_rows))
