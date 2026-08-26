#!/usr/bin/env python3
"""
build-draft-board.py - turn research-output/dataset.json into the Nasties draft
board (.xlsx), the printable sheet Joe drafts off of.

This is the committed, year-aware version of three throwaway scripts used in the
2026 prep (extract.mjs + sources.mjs + build_xlsx.py). It does the whole board
step in one command, with zero fabrication:

  research-output/dataset.json      (the app pipeline output - real engine numbers)
  research-output/source-counts.json (hand-tallied breakout/bust mentions - the ONE
                                      manual research input, edited fresh each year)
        -> per-position rows (ECR, tier, the 3 price columns, Target 1-5, pocket)
        -> merge in outlet counts by name match
        -> {League}_Draft_Board_{data-date}.xlsx

Every price/tier/pocket number comes straight from dataset.json. The only
web-researched fields are the two source-count columns, which come straight from
source-counts.json. Nothing is invented here.

Run:  python research-output/build-draft-board.py
      (from the fantasy_football_draft_app repo root)

Requires: openpyxl  (pip install openpyxl)
Copy rule (Joe): plain English, NO em/en dashes anywhere in surfaced strings.
"""
import json, os, re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- locate inputs relative to this file, so it runs from anywhere ------------
HERE = os.path.dirname(os.path.abspath(__file__))
DATASET = os.path.join(HERE, "dataset.json")
SOURCE_COUNTS = os.path.join(HERE, "source-counts.json")

def die(msg):
    print("ERROR:", msg)
    sys.exit(1)

if not os.path.exists(DATASET):
    die(f"{DATASET} not found. Run `npm run research:run` first to generate it.")
if not os.path.exists(SOURCE_COUNTS):
    die(f"{SOURCE_COUNTS} not found. It ships with the repo - restore it before building.")

with open(DATASET, encoding="utf-8") as f:
    ds = json.load(f)
players = ds if isinstance(ds, list) else (ds.get("players") or ds.get("rankings")
          or next((v for v in ds.values() if isinstance(v, list)), []))
if not players:
    die("dataset.json has no players array.")

meta = ds.get("meta", {}) if isinstance(ds, dict) else {}
league_name = meta.get("league", "Draft")
# data date drives the filename, so the board is stamped to the data, not today
gen = str(meta.get("generatedAt", ""))[:10] or "undated"
league_slug = re.sub(r"[^A-Za-z0-9]", "", league_name.replace("The ", "").split(" ")[0]) or "Draft"

with open(SOURCE_COUNTS, encoding="utf-8") as f:
    sc = json.load(f)
BREAKOUT_SLEEPER = sc.get("breakoutSleeper", {})
BUST = sc.get("bust", {})
ALIAS = sc.get("aliases", {})
SC_NOTE = sc.get("notes", "")

# --- Target 1-5 (ported verbatim from extract.mjs targetRating) ---------------
# Pocket (valueGap) is the spine: room-overpays => low target no matter how elite.
# Weighted by tier so a real pocket on a stud beats the same pocket on a scrub.
# Unavailable (OUT/IR/PUP/Doubtful) or "tax" capped low.
def target_rating(p):
    tag_ids = [t.get("id") for t in (p.get("tags") or [])]
    tag_labels = [(t.get("label") or "").upper() for t in (p.get("tags") or [])]
    tier = p["expertTier"] if isinstance(p.get("expertTier"), (int, float)) else 5
    gap = p["valueGap"] if isinstance(p.get("valueGap"), (int, float)) else 0
    w = 1.5 if tier <= 2 else 1.2 if tier <= 4 else 1.0 if tier <= 6 else 0.8
    adj = gap * w
    t = 5 if adj >= 8 else 4 if adj >= 4 else 3 if adj >= 1.5 else 2 if adj > -1 else 1
    if tier <= 2 and gap >= 1:
        t = min(5, t + 1)                       # landable elite bump
    if "tax" in tag_ids:
        t = min(t, 2)
    unavailable = ("out" in tag_ids) or any(
        re.search(r"\b(IR|PUP|DOUBTFUL|OUT)\b", l) for l in tag_labels)
    if unavailable:
        t = min(t, 2)
    return t

# --- name normalization for source-count matching (ported from sources.mjs) ---
def norm(n):
    s = n.lower()
    s = re.sub(r"[.’']", "", s)
    s = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return ALIAS.get(s, s)

def tally(outlet_map):
    t = {}
    for outlet, names in outlet_map.items():
        for n in names:
            t.setdefault(norm(n), set()).add(outlet)
    return t

bs_tally = tally(BREAKOUT_SLEEPER)
bust_tally = tally(BUST)

# --- build per-position rows --------------------------------------------------
POS = [("QB", "QB"), ("RB", "RB"), ("WR", "WR"), ("TE", "TE"), ("DEF", "Def")]
cols = {}
for code, label in POS:
    rows = []
    for p in players:
        if p.get("position") != code:
            continue
        band = p.get("valueBand") or {}
        k = norm(p.get("name", ""))
        rows.append({
            "ecr": p["ecrPositionRank"] if isinstance(p.get("ecrPositionRank"), (int, float)) else 9999,
            "name": p.get("name", ""),
            "position": label,
            "team": p.get("team", "") or "",
            "tier": p["expertTier"] if isinstance(p.get("expertTier"), (int, float)) else "",
            "goodPrice": band.get("low", ""),
            "winBid": band.get("base", ""),
            "maxPrice": band.get("high", ""),
            "target": target_rating(p),
            "roomPrice": p["expectedRoomPrice"] if isinstance(p.get("expectedRoomPrice"), (int, float)) else "",
            "pocket": p["valueGap"] if isinstance(p.get("valueGap"), (int, float)) else "",
            "breakoutSleeperSources": len(bs_tally[k]) if k in bs_tally else 0,
            "bustSources": len(bust_tally[k]) if k in bust_tally else 0,
            "injury": p.get("injuryStatus", "") or "",
        })
    rows.sort(key=lambda r: (r["ecr"], -r["target"]))
    for r in rows:
        if r["ecr"] == 9999:
            r["ecr"] = ""
    cols[label] = rows

# --- render the workbook (styling ported from build_xlsx.py) ------------------
HEAD = ["ECR", "Player", "Pos", "Team", "Tier", "Good $ (steal)", "Win $ (target bid)",
        "Max $ (walk-away)", "Target (1-5)", "Room $", "Pocket +/-",
        "Breakout/Sleeper Sources", "Bust Sources", "Injury"]
KEYS = ["ecr", "name", "position", "team", "tier", "goodPrice", "winBid", "maxPrice",
        "target", "roomPrice", "pocket", "breakoutSleeperSources", "bustSources", "injury"]
WIDTHS = [6, 24, 6, 7, 6, 13, 16, 16, 12, 8, 11, 26, 14, 13]

NAVY = "1F2A44"; GOLD = "C8A24B"; WHITE = "FFFFFF"
hdr_fill = PatternFill("solid", fgColor=NAVY)
hdr_font = Font(bold=True, color=WHITE, size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
TGT_FILL = {5: "C6EFCE", 4: "E2EFDA", 3: "FFFFFF", 2: "FCE4D6", 1: "FFC7CE"}
WIN_FILL = PatternFill("solid", fgColor="FFF2CC")  # highlight the target-bid column
WIN_COL, TGT_COL, POCKET_COL, SRC_COLS = 7, 9, 11, (12, 13)

def pocket_font(v):
    if isinstance(v, (int, float)):
        if v > 0: return Font(color="107C10")
        if v < 0: return Font(color="C00000")
    return Font()

wb = Workbook()
wb.remove(wb.active)

for _, title in POS:
    ws = wb.create_sheet(title)
    ws.append(HEAD)
    for ci, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 30
    for r in cols[title]:
        ws.append([r.get(k, "") if r.get(k, "") != "" else "" for k in KEYS])
        row = ws.max_row
        for ci in range(1, len(HEAD) + 1):
            cell = ws.cell(row=row, column=ci)
            cell.border = border
            cell.alignment = left if ci == 2 else center
        wc = ws.cell(row=row, column=WIN_COL); wc.fill = WIN_FILL; wc.font = Font(bold=True)
        tv = ws.cell(row=row, column=TGT_COL).value
        if tv in TGT_FILL:
            f = ws.cell(row=row, column=TGT_COL); f.fill = PatternFill("solid", fgColor=TGT_FILL[tv]); f.font = Font(bold=True)
        pc = ws.cell(row=row, column=POCKET_COL); pc.font = pocket_font(pc.value)
        for col_i in SRC_COLS:
            sc_cell = ws.cell(row=row, column=col_i)
            if isinstance(sc_cell.value, (int, float)) and sc_cell.value >= 2:
                sc_cell.font = Font(bold=True, color="9C27B0" if col_i == 12 else "C00000")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{ws.max_row}"

# --- README tab (year-aware; outlet lists pulled from source-counts.json) -----
bs_outlets = list(BREAKOUT_SLEEPER.keys())
bust_outlets = list(BUST.keys())
rd = wb.create_sheet("How to read this", 0)
rd.column_dimensions["A"].width = 120
rd.sheet_view.showGridLines = False
lines = [
 (f"THE {league_slug.upper()} - DRAFT BOARD (data generated {gen})", True, 14, GOLD),
 ("12-team | $200 | full-PPR | no-kicker | ESPN auction. One tab per position, sorted by ECR.", False, 11, None),
 ("", False, 11, None),
 ("THE THREE PRICE COLUMNS (this is the part that trips people up - read once)", True, 12, NAVY),
 ("Good $ (steal)     = bottom of the value band. Land him here and you clearly won the bid.", False, 11, None),
 ("Win $ (target bid) = the disciplined number you AIM for. Bid up to here and you got him at a good price. (highlighted yellow)", False, 11, None),
 ("Max $ (walk-away)  = the hard ceiling. The most he is worth even in the best case. Walk away one dollar over. Never bid past it.", False, 11, None),
 ("   The three numbers are the same value band: low / base / high from dataset.json. Aim for Win, stop at Max.", False, 11, None),
 ("", False, 11, None),
 ("THE REST OF THE COLUMNS", True, 12, NAVY),
 ("ECR = positional expert-consensus rank (FantasyPros, from the fresh pull). Player / Pos / Team / Tier (1 = elite).", False, 11, None),
 ("Room $ = what the room is projected to actually pay. Pocket +/- = worth minus room price (green = bargain, red = the room overpays).", False, 11, None),
 ("Target (1-5), 5 = highest-priority buy. Derived from the pocket edge weighted by tier: a real discount on a stud scores highest;", False, 11, None),
 ("   a player the room overpays for (negative pocket) scores 1 no matter how talented; OUT/IR/PUP/Doubtful and 'tax' tags are capped low.", False, 11, None),
 (f"Breakout/Sleeper Sources = how many separate outlets (of {len(bs_outlets)} checked) named him a breakout OR sleeper this year.", False, 11, None),
 (f"Bust Sources = how many separate outlets (of {len(bust_outlets)} checked) named him a bust / do-not-draft this year.", False, 11, None),
 ("   Read these two together: a high number = pervasive industry view; a 1 = one outlet making noise. A player with BOTH is polarizing.", False, 11, None),
 ("   Blank/0 = not named by any checked outlet (most of the pool).", False, 11, None),
 ("", False, 11, None),
 ("SOURCE OUTLETS (real articles, tallied by hand in source-counts.json - every count is traceable)", True, 12, NAVY),
 (f"Breakout/Sleeper ({len(bs_outlets)}): {', '.join(bs_outlets)}.", False, 11, None),
 (f"Bust ({len(bust_outlets)}): {', '.join(bust_outlets)}.", False, 11, None),
]
if SC_NOTE:
    lines.append((f"Note: {SC_NOTE}", False, 11, None))
lines += [
 ("", False, 11, None),
 ("Every column except the two source counts comes straight from research-output/dataset.json (the same file behind the target sheet).", False, 11, None),
 ("The source counts are the only web-researched fields; they are NOT in the model, they were gathered from the outlets above.", False, 11, None),
]
for i, (txt, bold, size, color) in enumerate(lines, 1):
    c = rd.cell(row=i, column=1, value=txt)
    c.font = Font(bold=bold, size=size, color=color if color else "000000")
    c.alignment = Alignment(vertical="center", wrap_text=False)

# --- save: a versioned copy in the repo, plus Downloads for draft day ---------
fname = f"{league_slug}_Draft_Board_{gen}.xlsx"
targets = [os.path.join(HERE, fname)]
downloads = os.path.join(os.path.expanduser("~"), "Downloads")
if os.path.isdir(downloads):
    targets.append(os.path.join(downloads, fname))

saved = []
for path in targets:
    wb.save(path)
    saved.append(path)

print("BUILT:", fname)
for path in saved:
    print("  saved ->", path, f"({os.path.getsize(path)} bytes)")
for _, title in POS:
    print(f"  {title}: {len(cols[title])} rows")
matched = sum(1 for pos in cols for r in cols[pos] if r["breakoutSleeperSources"] or r["bustSources"])
print(f"  players with >=1 source tag: {matched}")
